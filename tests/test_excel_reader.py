"""
DesktopAI
Tests for the Excel Reader module.
"""

import openpyxl

from documents.excel_reader import read_excel_text


def test_read_excel_text_extracts_cell_values(tmp_path):
    """A valid Excel file's cell values should be extracted correctly."""
    excel_path = tmp_path / "sample.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Score"
    sheet["A2"] = "Alice"
    sheet["B2"] = 95
    workbook.save(excel_path)

    result = read_excel_text(excel_path)

    assert result is not None
    assert "Name" in result
    assert "Alice" in result
    assert "95" in result


def test_read_excel_text_reads_all_sheets(tmp_path):
    """Values from every sheet should be extracted, not just the first."""
    excel_path = tmp_path / "multi_sheet.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "SheetOneValue"
    second_sheet = workbook.create_sheet("Sheet2")
    second_sheet["A1"] = "SheetTwoValue"
    workbook.save(excel_path)

    result = read_excel_text(excel_path)

    assert result is not None
    assert "SheetOneValue" in result
    assert "SheetTwoValue" in result


def test_read_excel_text_returns_none_for_missing_file(tmp_path):
    """A path that doesn't exist should return None, not raise."""
    missing_path = tmp_path / "does_not_exist.xlsx"

    result = read_excel_text(missing_path)

    assert result is None


def test_read_excel_text_returns_none_for_corrupt_file(tmp_path):
    """A file that isn't actually a valid Excel file should return
    None, not crash the caller."""
    fake_excel_path = tmp_path / "corrupt.xlsx"
    fake_excel_path.write_bytes(b"this is not a real xlsx")

    result = read_excel_text(fake_excel_path)

    assert result is None